from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.dimensions import SheetFormatProperties
import pandas as pd
import numpy as np
import warnings
import re

## non usare direttamente, la chiama il metodo "from_excel" della classe Portfolio_data!
def extract_data(worksheet, index = True):
    data = {}
    rows_list = []
    for row in worksheet.rows:
        row_as_list = []
        for cell in row:
            if cell.value == None:
                row_as_list.append('#N/A')
            else:
                row_as_list.append(cell.value)
        rows_list.append(row_as_list)
    rows_array = np.array(rows_list)
    rows_for_df = []
    row_len = rows_array.shape[1]
    na_row = np.array([['#N/A']*row_len])
    rows_array = np.vstack([na_row, rows_array, na_row])
    for i in range(1, rows_array.shape[0]):
        NAs = rows_array[i] == '#N/A'
        count_NAs = NAs.sum()
        if count_NAs == row_len:
            if (rows_array[i-1] == '#N/A').sum() == row_len:
                continue
            else:
                if len(rows_for_df) == 1:
                    rows_for_df_arr = np.array(rows_for_df)
                    cells_to_keep = rows_for_df_arr[0][rows_for_df_arr[0] != '#N/A']
                    if len(cells_to_keep) == 2:
                        data.update({cells_to_keep[0] : cells_to_keep[1]})
                    elif len(cells_to_keep) > 1:
                        data.update({cells_to_keep[0] : pd.Series(cells_to_keep[1:])})
                elif len(rows_for_df[0]) > 1:
                    rows_for_df_arr = np.array(rows_for_df)
                    while (rows_for_df_arr[:, 0] != '#N/A').sum() == 0:
                        rows_for_df_arr = np.delete(rows_for_df_arr, 0, axis = 1)
                    df = pd.DataFrame(rows_for_df_arr[1:], columns = rows_for_df_arr[0])
                    if index == True:
                        df.set_index(df.columns[0], inplace = True)
                    data.update({column : df[column] for column in df.columns if column != '#N/A'})
            rows_for_df = []
        else:
            rows_for_df.append(rows_array[i])
    return data



class Segment_data:
    def __init__(self, bank, ref_date, ws):
        self.bank = bank
        self.ref_date = ref_date
        data_dict = extract_data(ws)
        pattern = re.compile(r'\d')
        for key, value in data_dict.items():
            if type(value) == pd.Series:
                value = value.apply(lambda x: np.float64(x) if bool(pattern.match(x)) else np.nan)
            else:
                try:
                    value = np.float64(value)
                except:
                    pass
            setattr(self, key, value)
        self.EL_TTC = self.PD_CENTRAL*self.LGD
        try:
            self.STRESS_PARAM = self.PD_ADVERSE/self.PD_PIT
        except:
            self.STRESS_PARAM = np.nan
            warnings.warn('Insufficient data to compute stress parameter, try setting manually')
        try:
            self.EL_STRESS = self.EL_TTC*self.STRESS_PARAM
        except:
            self.EL_STRESS = np.nan
            warnings.warn('Insufficient data to compute stressed EL, try manually setting stress parameter')
        self.EL_THRESHOLDS = None
        self.CRA_COLOUR = None
    def calculate_CRA_COLOUR(self, one_threshold = False):
        EL_TTC = self.EL_TTC
        EL_STRESS = self.EL_STRESS
        absorbed_capital = self.RWA.sum()*self.CET1_LIMIT
        EL_thresholds_list = []
        EL_threshold = 0
        CRA_COLOUR = EL_STRESS.where(EL_STRESS < EL_threshold, 'Green')
        for i in range(2 - one_threshold):
            mask = EL_STRESS >= EL_threshold
            gross_profit = self.GROSS_PROFIT[mask].sum()
            exposure_onbal = self.EXPOSURE_ONBAL[mask].sum()
            op_cost = self.COST_INCOME*gross_profit
            cost_of_capital = (absorbed_capital/self.EXPOSURE.sum())*exposure_onbal*self.COC
            EL_threshold = (gross_profit - op_cost - (cost_of_capital/(1-self.TAX_RATE)))/exposure_onbal
            EL_thresholds_list.append(EL_threshold)
            if i == 0:
                CRA_COLOUR = CRA_COLOUR.where(EL_STRESS < EL_threshold, 'Yellow')
            if i == 1 and EL_thresholds_list[1] < EL_thresholds_list[0]:
                EL_threshold = EL_thresholds_list[0]
                print('Using only one threshold!')
        CRA_COLOUR = CRA_COLOUR.where(EL_TTC < EL_threshold, 'Red')
        self.CRA_COLOUR = CRA_COLOUR
        self.EL_THRESHOLDS = EL_thresholds_list
        return {'EL_thresholds' : EL_thresholds_list, 'COLOUR_SCALE' : CRA_COLOUR}
    def CRA_COLOUR_manual(self, rating_class, colour):
        if colour.title() not in ['Green', 'Yellow', 'Red']:
            raise ValueError('Misspecified CRA colour!')
        elif type(self.CRA_COLOUR) != pd.Series:
            raise AttributeError('Generate colour scale using the calculate_CRA_COLOUR method!')
        else:
            self.CRA_COLOUR[rating_class] = colour.title()
    def EAD_t0(self):
        return self.EAD.sum()
    def Defaults_t1(self):
        return (self.PD_PIT_ONE_YEAR*self.EAD).sum()
    def Cures_t1(self):
        try:
            return self.NPE_STOCK*self.CURE_RATE.mean() #improve
        except:
            return 0
    def Closures_amort_t1(self):
        return self.EAD_MLT*self.CLOSURES_AMORT
    def EAD_t1_wo_new_business(self):
        return self.EAD_t0() - self.Defaults_t1() + self.Cures_t1() - self.Closures_amort_t1()
    def EAD_t1_w_new_business(self):
        return self.EAD.sum()*(1+ self.EAD_GROWTH_RATE)
    def EAD_t1_new_business(self):
        return self.EAD_t1_w_new_business() - self.EAD_t1_wo_new_business()
    def EAD_projection(self):
        return pd.Series({f.__name__ : f() for f in
        [self.EAD_t0, self.Defaults_t1, self.Cures_t1, self.Closures_amort_t1, self.EAD_t1_wo_new_business, self.EAD_t1_w_new_business, self.EAD_t1_new_business]})


class Portfolio_data:
    def __init__(self, bank=None, ref_date=None, source = None, segments = {}):
        self.bank = bank
        self.ref_date = ref_date
        self.source = source
        self.segments = segments
        self.clusters_limit = None
    @classmethod
    def from_excel(cls, path):
        ref_date = pd.Timestamp(path.split('_')[0])
        bank = path.split('_')[1]
        source = load_workbook(path, data_only = True)
        segments = {}
        for worksheet in source.worksheets:
            if 'Input' in worksheet.title:
                segment_name = worksheet.title.split(" ")[-1]
                segments.update({segment_name : Segment_data(bank, ref_date, worksheet)})
        return cls(bank, ref_date, source, segments)
    def get_clusters_limit(self, *limits, portfolio_level = False):
        df = pd.DataFrame()
        for segment_name, segment in self.segments.items():
            segment_df = pd.DataFrame({'EAD_t0' : segment.EAD, 'EL_TTC' : segment.EL_TTC, 'CRA_COLOUR' : segment.CRA_COLOUR})
            segment_df['SEGMENT'] = segment_name
            df = df.append(segment_df)
        df['LIMIT'] = np.repeat(False, len(df))
        for item in limits:
            limit_i = np.repeat(True, len(df))
            for key, value in item.items():
                try:
                    limit_ij = df[key].isin(value)
                except TypeError:
                    limit_ij = df[key] == value
                limit_i = limit_i & limit_ij
            df['LIMIT'] = np.where(limit_i, True, df['LIMIT'])
        df_limit = df.where(df['LIMIT'])
        df_nolimit = df.where(~df['LIMIT'])
        df_nolimit = df_nolimit.apply(lambda x: 'All' if x.name not in ['LIMIT', 'EL_TTC', 'EAD_t0'] else x)
        df_limit = df_limit.append(df_nolimit.dropna())
        clusters_limit = df_limit.groupby(['SEGMENT']*(1-portfolio_level) + ['CRA_COLOUR']).agg({'EAD_t0' : np.sum, 'EL_TTC' : np.average})
        self.clusters_limit = clusters_limit.sort_values('EL_TTC')
        return self
    def set_limits(self, k, starting_exposure_manual = None, new_business_t1_manual = None):
        if starting_exposure_manual == None:
            try:
                starting_exposure = self.clusters_limit['EAD_t0']
            except TypeError:
                raise TypeError('Set parameter manually or set limit clusters first using the get_clusters_limit method!')
        else:
            starting_exposure = starting_exposure_manual
        starting_allocation = starting_exposure/starting_exposure.sum()
        if new_business_t1_manual == None:
            try:
                new_business_t1 = sum([segment.EAD_t1_new_business() for segment in self.segments.values()])
            except TypeError:
                raise TypeError('Set parameter manually or set limit clusters first using the get_clusters_limit method!')
        else:
            new_business_t1 = new_business_t1_manual
        el = self.clusters_limit['EL_TTC']
        EL_mean = el.mean()
        EL_var = el.var()
        n = len(el)
        new_allocation = starting_allocation + k*(el-EL_mean)/(n*EL_var)
        while sum(new_allocation < 0) > 0:
            mask = new_allocation > 0
            EL_mean = el[mask].mean()
            EL_var = el[mask].var()
            n_hat = sum(mask)
            new_allocation = np.where(mask,
                starting_allocation + (sum(starting_allocation[~mask])/n_hat) + ((el - EL_mean)/(n_hat*EL_var))*(k + sum(starting_allocation[~mask]*(el[~mask]-EL_mean))),
                0)
        self.clusters_limit['LIMIT_t1'] = new_allocation*new_business_t1
        return self
    def export_results(self):
        wb = self.source
        for segment_name, segment in self.segments.items():
            wb.create_sheet(segment_name + '_CRA_colour')
            df = pd.DataFrame({'EAD' : segment.EAD, 'EL_TTC' : segment.EL_TTC, 'EL_STRESS' : segment.EL_STRESS, 'CRA_COLOUR' : segment.CRA_COLOUR})
            count = 0
            for row in dataframe_to_rows(df, index = True, header = True):
                if count == 1:
                    pass
                else:
                    wb[segment_name + '_CRA_colour'].append(row)
                count += 1
            wb.create_sheet(segment_name + '_EAD_projection')
            df = pd.DataFrame({'EAD_projection' : segment.EAD_projection()})
            count = 0
            for row in dataframe_to_rows(df, index = True, header = True):
                if count == 1:
                    pass
                else:
                    wb[segment_name + '_EAD_projection'].append(row)
                count += 1
        wb.create_sheet('Limits')
        wb['Limits'].column_dimensions.width = 20
        df = self.clusters_limit.reset_index()
        count = 0
        for row in dataframe_to_rows(df, index = False, header = True):
            if count == 1:
                pass
            else:
                wb['Limits'].append(row)
            count += 1
        for sheet in wb.worksheets:
            sheet.sheet_format = SheetFormatProperties(baseColWidth = 20, defaultColWidth = 20)
        wb.save(self.bank + '_CRA_results.xlsx')

######################## PROVE ########################################
test = Portfolio_data.from_excel('20190416_PBZ_CRA_input_data.xlsx')
#
test.segments['SME'].STRESS_PARAM = 1.3288
test.segments['SME'].EL_STRESS = test.segments['SME'].STRESS_PARAM*test.segments['SME'].EL_TTC

test.segments['Large_corporate'].STRESS_PARAM = 1.3288
test.segments['Large_corporate'].EL_STRESS = test.segments['Large_corporate'].STRESS_PARAM*test.segments['Large_corporate'].EL_TTC
#
for segment in test.segments:
     test.segments[segment].calculate_CRA_COLOUR()
#
# test.segments['SME'].CRA_COLOUR_manual('SML7', 'Yellow')
print(test.segments['SME'].CRA_COLOUR)
#print(test.segments['SME'].EAD_projection())
#df = test.get_clusters_limit({'CRA_COLOUR' : ['Red', 'Yellow']}).clusters_limit
# #
test.get_clusters_limit({'CRA_COLOUR' : ['Red', 'Yellow']}).set_limits(k = -0.01).export_results()


# #
# #
#
# print(test.clusters_limit.applymap(lambda x: int(x/1000) if x > 1 else x))
# #
# print(test.segments['SME'].Cures_t1())
# print(test.source)
#
